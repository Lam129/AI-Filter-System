from pymongo import MongoClient
from lib import connect
from openpyxl import load_workbook
import time

start = time.time()
# Connect to excel
wb = load_workbook("features_lower_clustered.xlsx")
ws = wb.active
# Create array to store keywords
features_functionality = []
features_body = []
features_other = []

# Read each row in each column in the excel and append to the array
for i in range(ws.max_row):
    feature_functionality = ws.cell(row=(i+2), column=1).value
    if feature_functionality != None:
        features_functionality.append(feature_functionality)
    else:
        continue
for i in range(ws.max_row):
    feature_body = ws.cell(row=(i + 2), column=2).value
    if feature_body != None:
        features_body.append(feature_body)
    else:
        continue
for i in range(ws.max_row):
    feature_other = ws.cell(row=(i + 2), column=3).value
    if feature_other != None:
        features_other.append(feature_other)
    else:
        continue

# Connect to database
client = MongoClient(connect(), 27017)
db_in = client['FYP_result']
col_in = db_in['FYP_2021_Comment_list']
db_out = client['FYP_result']
col_out_functionality = db_out['FYP_2021_result_functionality']
col_out_body = db_out['FYP_2021_result_body']
col_out_other = db_out['FYP_2021_result_other']
# Create array to store comments
values_functionality = []
values_body = []
values_other = []

# Classify the comments according to keywords in feature list
for x in col_in.find({}, {"_id": 0, "product": 0, "title": 0}):
    for feature1 in features_functionality:
        value1 = x.get("body")
        value1 = value1.lower()
        if feature1 in value1:
            values_functionality.append(value1)

    for feature2 in features_body:
        value2 = x.get("body")
        value2 = value2.lower()
        if feature2 in value2:
            values_body.append(value2)

    for feature3 in features_other:
        value3 = x.get("body")
        value3 = value3.lower()
        if feature3 in value3:
            values_other.append(value3)

# Remove duplicates
values_functionality = list(dict.fromkeys(values_functionality))
values_body = list(dict.fromkeys(values_body))
values_other = list(dict.fromkeys(values_other))

col_out_functionality.delete_many({})  # Initialize output storage
col_out_body.delete_many({})  # Initialize output storage
col_out_other.delete_many({})  # Initialize output storage

# Input data to corresponding collection
x = 0
for x in range(len(values_functionality)):
    data1 = {'body': values_functionality[x]}
    col_out_functionality.insert_one(data1)

y = 0
for y in range(len(values_body)):
    data2 = {'body': values_body[y]}
    col_out_body.insert_one(data2)

z = 0
for z in range(len(values_other)):
    data3 = {'body': values_other[z]}
    col_out_other.insert_one(data3)


end = time.time()
print(end-start)




